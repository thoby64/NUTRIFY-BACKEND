from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func, text
from typing import List, Optional
from app.core.database import get_db
from app.core.auth import require_admin, require_manager_or_admin, require_any_of_roles
from app.models.models import Food, FoodGroup, Nutrient, NutrientType, FoodNutrient
from app.schemas.schemas import (
    FoodSearchRequest,
    FoodSearchResponse,
    FoodWithNutrients,
    FoodCompleteResponse,
    NutrientGrouped,
    FoodNutrientWithDetails,
    FoodGroupResponse,
    NutrientWithTypeResponse,
)
import os
import tempfile
from pathlib import Path
from io import BytesIO
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(
    prefix="/api/v1",
    tags=["nutrition"],
)


# ============ Food Group Endpoints ============
@router.get("/food-groups")
def get_food_groups(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=1000),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    user = Depends(require_any_of_roles(["admin", "manager", "nutritionist", "editor"]))
):
    """Get all food groups with optional search and pagination - requires manager or admin role"""
    query = db.query(FoodGroup)
    
    if search:
        query = query.filter(
            or_(
                FoodGroup.name.ilike(f"%{search}%"),
                FoodGroup.description.ilike(f"%{search}%")
            )
        )
    
    total = query.count()
    groups = query.offset(skip).limit(limit).all()
    
    return {
        "total": total,
        "data": [
            {
                "id": g.id,
                "name": g.name,
                "description": g.description,
                "foods_count": db.query(Food).filter(Food.food_group_id == g.id).count()
            }
            for g in groups
        ]
    }


# ============ Nutrient Endpoints ============
@router.get("/nutrient-types")
def get_nutrient_types(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=1000),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    user = Depends(require_any_of_roles(["admin", "manager", "nutritionist", "editor"]))
):
    """Get all nutrient types with optional search and pagination - requires manager or admin role"""
    query = db.query(NutrientType)
    
    if search:
        query = query.filter(
            or_(
                NutrientType.name.ilike(f"%{search}%"),
                NutrientType.category.ilike(f"%{search}%")
            )
        )
    
    total = query.count()
    types = query.offset(skip).limit(limit).all()
    
    result = []
    for nt in types:
        nutrients = db.query(Nutrient).filter(Nutrient.nutrient_type_id == nt.id).all()
        result.append({
            "id": nt.id,
            "name": nt.name,
            "category": nt.category,
            "nutrients_count": len(nutrients)
        })
    return {
        "total": total,
        "data": result
    }


# ============ Food Search & Filter Endpoints ============
@router.post("/foods/search", response_model=FoodSearchResponse)
def search_foods(
    request: FoodSearchRequest,
    db: Session = Depends(get_db),
    user = Depends(require_any_of_roles(["admin", "manager", "nutritionist", "editor"])),
):
    """
    Search and filter foods by multiple nutrient conditions - requires admin role
    
    Example:
    {
        "food_group_id": 1,
        "nutrient_conditions": [
            {"nutrient_id": 1, "min_value": 10},
            {"nutrient_id": 5, "max_value": 200}
        ],
        "sort_by_nutrient_id": 1,
        "limit": 50
    }
    """
    query = db.query(Food)
    
    # Filter by food group
    if request.food_group_id:
        query = query.filter(Food.food_group_id == request.food_group_id)
    
    # Filter by food name
    if request.search_name:
        query = query.filter(Food.name.ilike(f"%{request.search_name}%"))
    
    # Apply nutrient conditions (complex filtering)
    if request.nutrient_conditions:
        for condition in request.nutrient_conditions:
            # Create subquery for each condition
            subquery = db.query(FoodNutrient.food_id).filter(
                FoodNutrient.nutrient_id == condition.nutrient_id
            )
            
            if condition.operator == "range":
                if condition.min_value is not None:
                    subquery = subquery.filter(FoodNutrient.value >= condition.min_value)
                if condition.max_value is not None:
                    subquery = subquery.filter(FoodNutrient.value <= condition.max_value)
            elif condition.operator == "min":
                subquery = subquery.filter(FoodNutrient.value >= condition.min_value)
            elif condition.operator == "max":
                subquery = subquery.filter(FoodNutrient.value <= condition.max_value)
            elif condition.operator == "equals":
                subquery = subquery.filter(FoodNutrient.value == condition.min_value)
            
            query = query.filter(Food.id.in_(subquery))
    
    # Get total count before pagination
    total_count = query.count()
    
    # Sorting by nutrient
    if request.sort_by_nutrient_id:
        nutrient_query = db.query(
            FoodNutrient.food_id,
            FoodNutrient.value
        ).filter(
            FoodNutrient.nutrient_id == request.sort_by_nutrient_id
        )
        
        food_ids = [fn.food_id for fn in nutrient_query.all()]
        
        if food_ids:
            # Order query based on nutrient values
            query = query.filter(Food.id.in_(food_ids))
            
            # Get values for sorting
            sorted_foods = []
            for food_id in food_ids:
                fn = db.query(FoodNutrient).filter(
                    and_(
                        FoodNutrient.food_id == food_id,
                        FoodNutrient.nutrient_id == request.sort_by_nutrient_id,
                    )
                ).first()
                if fn and fn.value is not None:
                    sorted_foods.append((food_id, fn.value))
            
            sorted_foods.sort(key=lambda x: x[1], reverse=(request.sort_order == "desc"))
            sorted_ids = [f[0] for f in sorted_foods]
            
            # Reorder query results
            from sqlalchemy import case
            order = case({id_: idx for idx, id_ in enumerate(sorted_ids)})
            query = query.order_by(order)
    
    # Pagination
    foods = query.offset(request.offset).limit(request.limit).all()
    
    # Build response with nutrient details
    food_responses = []
    for food in foods:
        # Group nutrients by type
        nutrient_groups = []
        nutrient_types = db.query(NutrientType).all()
        
        for nt in nutrient_types:
            nutrients_in_type = db.query(
                FoodNutrient.id,
                FoodNutrient.value,
                FoodNutrient.per_unit,
                FoodNutrient.nutrient_id,
                FoodNutrient.nutrient_type_id,
                Nutrient.name.label("nutrient_name"),
                Nutrient.unit.label("nutrient_unit"),
                FoodNutrient.data_source,
                NutrientType.name.label("nutrient_type_name"),
            ).join(
                Nutrient, FoodNutrient.nutrient_id == Nutrient.id
            ).join(
                NutrientType, FoodNutrient.nutrient_type_id == NutrientType.id
            ).filter(
                and_(
                    FoodNutrient.food_id == food.id,
                    FoodNutrient.nutrient_type_id == nt.id,
                )
            ).all()
            
            if nutrients_in_type:
                nutrient_group = NutrientGrouped(
                    nutrient_type_id=nt.id,
                    nutrient_type_name=nt.name,
                    nutrients=[
                        FoodNutrientWithDetails(
                            id=n.id,
                            value=n.value,
                            per_unit=n.per_unit,
                            nutrient_id=n.nutrient_id,
                            nutrient_type_id=n.nutrient_type_id,
                            nutrient_name=n.nutrient_name,
                            nutrient_unit=n.nutrient_unit,
                            nutrient_type_name=n.nutrient_type_name,
                            data_source=n.data_source,
                        )
                        for n in nutrients_in_type
                    ]
                )
                nutrient_groups.append(nutrient_group)
        
        food_group = db.query(FoodGroup).filter(FoodGroup.id == food.food_group_id).first()
        
        food_response = FoodCompleteResponse(
            id=food.id,
            name=food.name,
            code=food.code,
            description=food.description,
            food_group_id=food.food_group_id,
            food_group_name=food_group.name if food_group else "",
            nutrient_groups=nutrient_groups,
            created_at=food.created_at,
        )
        food_responses.append(food_response)
    
    return FoodSearchResponse(
        total_count=total_count,
        limit=request.limit,
        offset=request.offset,
        foods=food_responses,
    )


@router.get("/foods/{food_id}", response_model=FoodCompleteResponse)
def get_food_details(
    food_id: int,
    db: Session = Depends(get_db),
    user = Depends(require_any_of_roles(["admin", "manager", "nutritionist", "editor"])),
):
    """Get complete details for a single food - requires manager or admin role"""
    food = db.query(Food).filter(Food.id == food_id).first()
    
    if not food:
        raise HTTPException(status_code=404, detail="Food not found")
    
    # Group nutrients by type
    nutrient_groups = []
    nutrient_types = db.query(NutrientType).all()
    
    for nt in nutrient_types:
        nutrients_in_type = db.query(
            FoodNutrient.id,
            FoodNutrient.value,
            FoodNutrient.per_unit,
            FoodNutrient.nutrient_id,
            FoodNutrient.nutrient_type_id,
            Nutrient.name.label("nutrient_name"),
            Nutrient.unit.label("nutrient_unit"),
            FoodNutrient.data_source,
            NutrientType.name.label("nutrient_type_name"),
        ).join(
            Nutrient, FoodNutrient.nutrient_id == Nutrient.id
        ).join(
            NutrientType, FoodNutrient.nutrient_type_id == NutrientType.id
        ).filter(
            and_(
                FoodNutrient.food_id == food_id,
                FoodNutrient.nutrient_type_id == nt.id,
            )
        ).all()
        
        # Add nutrient group even if empty
        nutrient_group = NutrientGrouped(
            nutrient_type_id=nt.id,
            nutrient_type_name=nt.name,
            nutrients=[
                FoodNutrientWithDetails(
                    id=n.id,
                    value=n.value,
                    per_unit=n.per_unit,
                    nutrient_id=n.nutrient_id,
                    nutrient_type_id=n.nutrient_type_id,
                    nutrient_name=n.nutrient_name,
                    nutrient_unit=n.nutrient_unit,
                    nutrient_type_name=n.nutrient_type_name,
                    data_source=n.data_source,
                )
                for n in nutrients_in_type
            ] if nutrients_in_type else []
        )
        nutrient_groups.append(nutrient_group)
    
    food_group = db.query(FoodGroup).filter(FoodGroup.id == food.food_group_id).first()
    
    return FoodCompleteResponse(
        id=food.id,
        name=food.name,
        code=food.code,
        description=food.description,
        food_group_id=food.food_group_id,
        food_group_name=food_group.name if food_group else "",
        nutrient_groups=nutrient_groups,
        created_at=food.created_at,
    )


@router.get("/foods")
def list_foods(
    food_group_id: Optional[int] = None,
    search: Optional[str] = None,
    limit: int = Query(50, le=1000),
    skip: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user = Depends(require_any_of_roles(["admin", "manager", "nutritionist", "editor"])),
):
    """List all foods with optional filtering - requires manager or admin role"""
    query = db.query(Food)
    
    if food_group_id:
        query = query.filter(Food.food_group_id == food_group_id)
    
    if search:
        query = query.filter(Food.name.ilike(f"%{search}%"))
    
    total = query.count()
    foods = query.offset(skip).limit(limit).all()
    
    # Build response
    result = []
    for food in foods:
        food_group = db.query(FoodGroup).filter(FoodGroup.id == food.food_group_id).first()
        nutrients_count = db.query(FoodNutrient).filter(FoodNutrient.food_id == food.id).count()
        result.append({
            "id": food.id,
            "name": food.name,
            "code": food.code,
            "description": food.description,
            "food_group_id": food.food_group_id,
            "food_group_name": food_group.name if food_group else None,
            "nutrients_count": nutrients_count,
            "created_at": food.created_at.isoformat() if food.created_at else None,
        })
    
    return {
        "total": total,
        "data": result
    }


# ============ Nutrient GET endpoints ============
@router.get("/nutrients", response_model=dict)
def get_nutrients(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=1000),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    user = Depends(require_any_of_roles(["admin", "manager", "nutritionist", "editor"])),
):
    """Get all nutrients with optional search - requires manager or admin role"""
    query = db.query(Nutrient)
    
    if search:
        query = query.filter(Nutrient.name.ilike(f"%{search}%"))
    
    total = query.count()
    nutrients = query.offset(skip).limit(limit).all()
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "data": [
            {
                "id": n.id,
                "name": n.name,
                "abbreviation": n.abbreviation,
                "unit": n.unit or (db.query(FoodNutrient).filter(FoodNutrient.nutrient_id == n.id).first().per_unit if db.query(FoodNutrient).filter(FoodNutrient.nutrient_id == n.id).first() else None),
                "nutrient_type_id": n.nutrient_type_id,
                "nutrient_type_name": db.query(NutrientType).filter(NutrientType.id == n.nutrient_type_id).first().name if n.nutrient_type_id else None,
                "nutrients_count": db.query(FoodNutrient).filter(FoodNutrient.nutrient_id == n.id).count(),
            }
            for n in nutrients
        ]
    }


@router.get("/food-nutrients", response_model=dict)
def get_food_nutrients(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=1000),
    nutrient_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user = Depends(require_any_of_roles(["admin", "manager", "nutritionist", "editor"])),
):
    """Get all food-nutrient relationships - requires manager or admin role"""
    query = db.query(FoodNutrient)
    
    if nutrient_id:
        query = query.filter(FoodNutrient.nutrient_id == nutrient_id)
    
    total = query.count()
    food_nutrients = query.offset(skip).limit(limit).all()
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "data": [
            {
                "id": fn.id,
                "food_id": fn.food_id,
                "food_name": db.query(Food).filter(Food.id == fn.food_id).first().name if fn.food_id else None,
                "nutrient_id": fn.nutrient_id,
                "value": fn.value,
                "per_unit": fn.per_unit,
            }
            for fn in food_nutrients
        ]
    }


# ============ Import Endpoints ============
@router.post("/import-csv")
async def import_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin_user = Depends(require_admin)
):
    """
    Import a single CSV file into the database
    Requires admin role and CSV file format
    Frontend should send files one at a time
    """
    try:
        import tempfile
        import shutil

        if not file.filename.lower().endswith('.csv'):
            raise HTTPException(
                status_code=400,
                detail="Only CSV files are supported. Please upload a CSV file."
            )

        temp_dir = tempfile.mkdtemp()
        temp_file_path = Path(temp_dir) / file.filename
        content = await file.read()
        with open(temp_file_path, 'wb') as f:
            f.write(content)

        from import_data import DataImporter
        importer = DataImporter(temp_dir, db)
        success = importer.import_file(file.filename)

        shutil.rmtree(temp_dir)

        if not success:
            raise HTTPException(
                status_code=400,
                detail=f"Failed to import file: {file.filename}. Check file format and content."
            )

        return {
            "success": True,
            "imported_count": 1,
            "filename": file.filename,
            "message": f"Successfully imported {file.filename}."
        }
    except HTTPException:
        raise
    except Exception as e:
        if 'temp_dir' in locals():
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
        raise HTTPException(
            status_code=400,
            detail=f"Error importing CSV file: {str(e)}"
        )


@router.get("/templates/{filename}")
def get_template(
    filename: str,
    admin_user = Depends(require_admin)
):
    """
    Serve a sample template file
    Supported files: amincereals_090405.csv, macrocereals_091101.csv, etc.
    """
    # Path to CSV files in project root
    base_path = Path(__file__).parent.parent.parent.parent
    csv_path = base_path / filename
    
    # Security check: ensure file is in root directory and is a CSV
    if not filename.lower().endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are supported")
    
    if not csv_path.parent == base_path:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if not csv_path.exists():
        raise HTTPException(status_code=404, detail=f"Template file '{filename}' not found")
    
    # Return file as download
    from fastapi.responses import FileResponse
    return FileResponse(
        path=csv_path,
        filename=filename,
        media_type='text/csv',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/reset-database")
def reset_database(
    db: Session = Depends(get_db),
    admin_user = Depends(require_admin)
):
    """
    Delete all nutrition data from the database (except users and system tables)
    This allows starting fresh with new imports
    Requires admin role for security
    """
    try:
        # Get counts before deletion
        food_nutrient_count = db.query(FoodNutrient).count()
        food_count = db.query(Food).count()
        nutrient_count = db.query(Nutrient).count()
        food_group_count = db.query(FoodGroup).count()
        nutrient_type_count = db.query(NutrientType).count()
        
        # Delete in correct order (respecting foreign key constraints)
        # 1. Delete all food-nutrient relationships first
        db.query(FoodNutrient).delete()
        db.commit()
        
        # 2. Delete all foods
        db.query(Food).delete()
        db.commit()
        
        # 3. Delete all nutrients
        db.query(Nutrient).delete()
        db.commit()
        
        # 4. Delete all food groups
        db.query(FoodGroup).delete()
        db.commit()
        
        # 5. Delete all nutrient types
        db.query(NutrientType).delete()
        db.commit()
        
        return {
            "success": True,
            "message": "Database reset successfully",
            "deleted_counts": {
                "food_nutrients": food_nutrient_count,
                "foods": food_count,
                "nutrients": nutrient_count,
                "food_groups": food_group_count,
                "nutrient_types": nutrient_type_count,
            },
            "total_deleted": food_nutrient_count + food_count + nutrient_count + food_group_count + nutrient_type_count
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Error resetting database: {str(e)}"
        )


# ============ Export Endpoints ============
@router.get("/export-excel")
def export_excel(
    db: Session = Depends(get_db),
    admin_user = Depends(require_admin)
):
    """
    Export all nutrition data to Excel file with multiple sheets
    Includes: Foods, Nutrients, Food Groups, Nutrient Types, Food-Nutrient Values
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=400,
            detail="Excel export not available. openpyxl is not installed."
        )
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 1. Food Groups Sheet
        ws_groups = wb.create_sheet("Food Groups")
        groups = db.query(FoodGroup).all()
        ws_groups.append(["ID", "Name", "Description"])
        for group in groups:
            ws_groups.append([group.id, group.name, group.description or ""])
        
        # Style headers
        for cell in ws_groups[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        ws_groups.column_dimensions['A'].width = 10
        ws_groups.column_dimensions['B'].width = 25
        ws_groups.column_dimensions['C'].width = 40
        
        # 2. Nutrient Types Sheet
        ws_types = wb.create_sheet("Nutrient Types")
        types = db.query(NutrientType).all()
        ws_types.append(["ID", "Name", "Category", "Description"])
        for nt in types:
            ws_types.append([nt.id, nt.name, nt.category, nt.description or ""])
        
        for cell in ws_types[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        ws_types.column_dimensions['A'].width = 10
        ws_types.column_dimensions['B'].width = 20
        ws_types.column_dimensions['C'].width = 20
        ws_types.column_dimensions['D'].width = 40
        
        # 3. Nutrients Sheet
        ws_nutrients = wb.create_sheet("Nutrients")
        nutrients = db.query(Nutrient).all()
        ws_nutrients.append(["ID", "Name", "Nutrient Type", "Unit", "Abbreviation"])
        for nutrient in nutrients:
            nt_name = db.query(NutrientType).filter(NutrientType.id == nutrient.nutrient_type_id).first()
            ws_nutrients.append([
                nutrient.id,
                nutrient.name,
                nt_name.name if nt_name else "",
                nutrient.unit or "",
                nutrient.abbreviation or ""
            ])
        
        for cell in ws_nutrients[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        ws_nutrients.column_dimensions['A'].width = 10
        ws_nutrients.column_dimensions['B'].width = 25
        ws_nutrients.column_dimensions['C'].width = 20
        ws_nutrients.column_dimensions['D'].width = 12
        ws_nutrients.column_dimensions['E'].width = 15
        
        # 4. Foods Sheet
        ws_foods = wb.create_sheet("Foods")
        foods = db.query(Food).all()
        ws_foods.append(["ID", "Code", "Name", "Food Group", "Created At"])
        for food in foods:
            fg = db.query(FoodGroup).filter(FoodGroup.id == food.food_group_id).first()
            ws_foods.append([
                food.id,
                food.code or "",
                food.name,
                fg.name if fg else "",
                food.created_at.strftime('%Y-%m-%d %H:%M:%S') if food.created_at else ""
            ])
        
        for cell in ws_foods[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        ws_foods.column_dimensions['A'].width = 10
        ws_foods.column_dimensions['B'].width = 15
        ws_foods.column_dimensions['C'].width = 30
        ws_foods.column_dimensions['D'].width = 25
        ws_foods.column_dimensions['E'].width = 20
        
        # 5. Food-Nutrient Values Sheet
        ws_values = wb.create_sheet("Nutrient Values")
        values = db.query(FoodNutrient).all()
        ws_values.append(["ID", "Food ID", "Food Name", "Nutrient ID", "Nutrient", "Value", "Unit", "Type"])
        for val in values:
            food = db.query(Food).filter(Food.id == val.food_id).first()
            nutrient = db.query(Nutrient).filter(Nutrient.id == val.nutrient_id).first()
            nt = db.query(NutrientType).filter(NutrientType.id == val.nutrient_type_id).first()
            
            ws_values.append([
                val.id,
                val.food_id,
                food.name if food else "",
                val.nutrient_id,
                nutrient.name if nutrient else "",
                val.value,
                val.per_unit or "",
                nt.name if nt else ""
            ])
        
        for cell in ws_values[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        ws_values.column_dimensions['A'].width = 10
        ws_values.column_dimensions['B'].width = 10
        ws_values.column_dimensions['C'].width = 25
        ws_values.column_dimensions['D'].width = 12
        ws_values.column_dimensions['E'].width = 25
        ws_values.column_dimensions['F'].width = 12
        ws_values.column_dimensions['G'].width = 10
        ws_values.column_dimensions['H'].width = 20
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"nutrition_data_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error exporting to Excel: {str(e)}"
        )


@router.get("/export-sql")
def export_sql(
    db: Session = Depends(get_db),
    admin_user = Depends(require_admin)
):
    """
    Export all nutrition data as SQL INSERT statements
    Useful for database backup and migration
    Schema matches the actual database tables exactly
    """
    try:
        sql_lines = [
            "-- ============================================================================",
            "-- Nutrition Database Export",
            f"-- Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "-- DO NOT MODIFY UNLESS YOU KNOW THE SCHEMA",
            "-- ============================================================================",
            ""
        ]
        
        # 1. Food Groups (food_groups table)
        sql_lines.append("-- ============================================================================")
        sql_lines.append("-- Food Groups Table")
        sql_lines.append("-- ============================================================================")
        groups = db.query(FoodGroup).all()
        for group in groups:
            desc = group.description.replace("'", "''") if group.description else ""
            created = group.created_at.strftime('%Y-%m-%d %H:%M:%S') if group.created_at else "NULL"
            sql_lines.append(
                f"INSERT INTO food_groups (id, name, description, created_at) VALUES ({group.id}, '{group.name.replace(chr(39), chr(39)+chr(39))}', '{desc}', '{created}');"
            )
        sql_lines.append("")
        
        # 2. Nutrient Types (nutrient_types table) - MUST EXPORT ALL TYPES FIRST
        sql_lines.append("-- ============================================================================")
        sql_lines.append("-- Nutrient Types Table (required by Nutrients)")
        sql_lines.append("-- ============================================================================")
        
        # First, find ALL nutrient_type_ids that are referenced in the nutrients table
        nutrients_query = db.query(Nutrient).all()
        referenced_nutrient_type_ids = set([n.nutrient_type_id for n in nutrients_query])
        
        # Export nutrient_types that are actually referenced OR exist in the table
        types = db.query(NutrientType).all()
        exported_nutrient_type_ids = set()
        
        for nt in types:
            if nt.id in referenced_nutrient_type_ids:
                exported_nutrient_type_ids.add(nt.id)
                desc = nt.description.replace("'", "''") if nt.description else ""
                cat = nt.category.replace("'", "''") if nt.category else "NULL"
                cat = f"'{cat}'" if nt.category else "NULL"
                created = nt.created_at.strftime('%Y-%m-%d %H:%M:%S') if nt.created_at else "NULL"
                sql_lines.append(
                    f"INSERT INTO nutrient_types (id, name, category, description, created_at) VALUES ({nt.id}, '{nt.name.replace(chr(39), chr(39)+chr(39))}', {cat}, '{desc}', '{created}');"
                )
        
        # Log which nutrient_type_ids are referenced but don't exist in the table
        missing_types = referenced_nutrient_type_ids - set([nt.id for nt in types])
        if missing_types:
            sql_lines.append(f"-- WARNING: The following nutrient_type_ids are referenced in nutrients but don't exist in nutrient_types: {missing_types}")
        
        sql_lines.append("")
        
        # 3. Nutrients (nutrients table) - ONLY EXPORT NUTRIENTS WITH VALID NUTRIENT_TYPE_ID
        sql_lines.append("-- ============================================================================")
        sql_lines.append("-- Nutrients Table (only with valid nutrient_type references)")
        sql_lines.append("-- ============================================================================")
        nutrients = nutrients_query  # Use the query result we already have
        exported_nutrient_ids = set()  # Track which nutrients are actually exported
        skipped_nutrients = 0
        for nutrient in nutrients:
            # Skip nutrients with non-existent nutrient_type references
            if nutrient.nutrient_type_id not in exported_nutrient_type_ids:
                sql_lines.append(f"-- SKIPPED NUTRIENT: '{nutrient.name}' references non-existent nutrient_type_id {nutrient.nutrient_type_id}")
                skipped_nutrients += 1
                continue
            
            # This nutrient is valid and will be exported
            exported_nutrient_ids.add(nutrient.id)
            
            unit = nutrient.unit.replace("'", "''") if nutrient.unit else "NULL"
            unit = f"'{unit}'" if nutrient.unit else "NULL"
            abbr = nutrient.abbreviation.replace("'", "''") if nutrient.abbreviation else "NULL"
            abbr = f"'{abbr}'" if nutrient.abbreviation else "NULL"
            desc = nutrient.description.replace("'", "''") if nutrient.description else ""
            created = nutrient.created_at.strftime('%Y-%m-%d %H:%M:%S') if nutrient.created_at else "NULL"
            sql_lines.append(
                f"INSERT INTO nutrients (id, nutrient_type_id, name, unit, abbreviation, description, created_at) VALUES ({nutrient.id}, {nutrient.nutrient_type_id}, '{nutrient.name.replace(chr(39), chr(39)+chr(39))}', {unit}, {abbr}, '{desc}', '{created}');"
            )
        sql_lines.append("")
        
        # 4. Foods (foods table) - ONLY EXPORT FOODS WITH VALID FOOD_GROUP_ID
        sql_lines.append("-- ============================================================================")
        sql_lines.append("-- Foods Table (only with valid food_group references)")
        sql_lines.append("-- ============================================================================")
        exported_food_group_ids = set([g.id for g in groups])
        foods = db.query(Food).all()
        exported_food_ids = set()
        for food in foods:
            # Skip foods with non-existent food_group references
            if food.food_group_id not in exported_food_group_ids:
                sql_lines.append(f"-- SKIPPED: Food '{food.name}' references non-existent food_group_id {food.food_group_id}")
                continue
            
            exported_food_ids.add(food.id)
            code = food.code.replace("'", "''") if food.code else "NULL"
            code = f"'{code}'" if food.code else "NULL"
            desc = food.description.replace("'", "''") if food.description else ""
            created = food.created_at.strftime('%Y-%m-%d %H:%M:%S') if food.created_at else "NULL"
            sql_lines.append(
                f"INSERT INTO foods (id, food_group_id, name, code, description, created_at) VALUES ({food.id}, {food.food_group_id}, '{food.name.replace(chr(39), chr(39)+chr(39))}', {code}, '{desc}', '{created}');"
            )
        sql_lines.append("")
        
        # 5. Food-Nutrient Values (food_nutrients table) - ONLY EXPORT VALID REFERENCES
        sql_lines.append("-- ============================================================================")
        sql_lines.append("-- Food-Nutrient Values Table (only with valid references)")
        sql_lines.append("-- ============================================================================")
        values = db.query(FoodNutrient).all()
        skipped_count = 0
        for val in values:
            # Skip food_nutrients with invalid references
            if val.food_id not in exported_food_ids:
                skipped_count += 1
                continue
            if val.nutrient_id not in exported_nutrient_ids:
                skipped_count += 1
                continue
            if val.nutrient_type_id not in exported_nutrient_type_ids:
                skipped_count += 1
                continue
            
            unit = val.per_unit.replace("'", "''") if val.per_unit else "NULL"
            unit = f"'{unit}'" if val.per_unit else "NULL"
            source = val.data_source.replace("'", "''") if val.data_source else "NULL"
            source = f"'{source}'" if val.data_source else "NULL"
            created = val.created_at.strftime('%Y-%m-%d %H:%M:%S') if val.created_at else "NULL"
            sql_lines.append(
                f"INSERT INTO food_nutrients (id, food_id, nutrient_id, nutrient_type_id, value, per_unit, data_source, created_at) VALUES ({val.id}, {val.food_id}, {val.nutrient_id}, {val.nutrient_type_id}, {val.value}, {unit}, {source}, '{created}');"
            )
        
        sql_lines.append("")
        sql_lines.append("-- ============================================================================")
        sql_lines.append("-- Export Summary")
        sql_lines.append("-- ============================================================================")
        sql_lines.append(f"-- Food Groups exported: {len(groups)}")
        sql_lines.append(f"-- Nutrient Types exported: {len(exported_nutrient_type_ids)}")
        sql_lines.append(f"-- Nutrients exported: {len([n for n in nutrients if n.nutrient_type_id in exported_nutrient_type_ids])}")
        sql_lines.append(f"-- Nutrients skipped (orphaned): {skipped_nutrients}")
        sql_lines.append(f"-- Foods exported: {len(exported_food_ids)}")
        sql_lines.append(f"-- Food-Nutrient values exported: {len(values) - skipped_count}")
        if skipped_count > 0:
            sql_lines.append(f"-- WARNING: {skipped_count} food_nutrient records skipped due to broken references")
        sql_lines.append("-- This export includes ONLY valid data (no orphaned records)")
        sql_lines.append("-- ============================================================================")
        
        sql_content = "\n".join(sql_lines)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"nutrition_data_export_{timestamp}.sql"
        
        return StreamingResponse(
            iter([sql_content.encode()]),
            media_type="text/plain",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error exporting to SQL: {str(e)}"
        )


@router.post("/import-sql")
async def import_sql(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin_user = Depends(require_admin)
):
    """
    Import SQL file with INSERT statements into the database
    Requires admin role and SQL file format
    Only executes INSERT statements, ignores CREATE TABLE statements
    Executes inserts in dependency order (food_groups → nutrient_types → nutrients → foods → food_nutrients)
    """
    # Validate file extension
    if not file.filename.lower().endswith('.sql'):
        raise HTTPException(
            status_code=400,
            detail="Only SQL files are supported. Please upload a .sql file."
        )
    
    try:
        # Read uploaded file
        content = await file.read()
        sql_text = content.decode('utf-8')
        
        # Parse SQL file to extract INSERT statements (sorted by dependency)
        insert_statements = parse_sql_inserts(sql_text)
        
        if not insert_statements:
            raise HTTPException(
                status_code=400,
                detail="No INSERT statements found in the SQL file"
            )
        
        # Execute INSERT statements
        executed_count = 0
        failed_count = 0
        errors = []
        
        for i, statement in enumerate(insert_statements):
            try:
                # Execute each statement separately
                db.execute(text(statement))
                db.commit()  # Commit after each successful insert
                executed_count += 1
            except Exception as e:
                db.rollback()  # Rollback only this statement
                failed_count += 1
                error_msg = str(e)[:150]
                errors.append(f"Statement {i+1}: {error_msg}")
        
        return {
            "success": executed_count > 0,
            "filename": file.filename,
            "inserted_count": executed_count,
            "failed_count": failed_count,
            "message": f"Successfully imported {executed_count} records from {file.filename}",
            "errors": errors[:10] if errors else []  # Return first 10 errors
        }
    
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Error importing SQL file: {str(e)[:200]}"
        )


def parse_sql_inserts(sql_text: str) -> List[str]:
    """
    Parse SQL file and extract INSERT statements
    Sorts them in correct order to respect foreign key constraints
    
    Args:
        sql_text: Raw SQL file content
    
    Returns:
        List of INSERT statements sorted by table dependency order
    """
    # Split by semicolon
    statements = sql_text.split(';')
    insert_statements = []
    
    for statement in statements:
        # Strip whitespace
        statement = statement.strip()
        
        # Skip empty statements
        if not statement:
            continue
        
        # Remove SQL comments from statement FIRST, before checking content
        lines = statement.split('\n')
        cleaned_lines = []
        for line in lines:
            # Remove inline comments
            if '--' in line:
                line = line[:line.index('--')]
            # Remove multi-line comment markers
            if line.strip() and not line.strip().startswith('/*'):
                cleaned_lines.append(line)
        
        cleaned_statement = '\n'.join(cleaned_lines).strip()
        
        # Skip empty statements after comment removal
        if not cleaned_statement:
            continue
        
        # Only include INSERT statements
        if cleaned_statement.upper().startswith('INSERT INTO'):
            # Add back the semicolon for execution
            insert_statements.append(cleaned_statement + ';')
    
    # Sort INSERT statements by table dependency order
    # This ensures foreign keys are satisfied
    table_order = ['food_groups', 'nutrient_types', 'nutrients', 'foods', 'food_nutrients']
    
    def get_table_name(statement: str) -> str:
        """Extract table name from INSERT INTO statement"""
        # Find "INSERT INTO" and get the next word
        upper_stmt = statement.upper()
        if 'INSERT INTO' not in upper_stmt:
            return ''
        
        start_idx = upper_stmt.index('INSERT INTO') + len('INSERT INTO')
        # Skip whitespace
        while start_idx < len(statement) and statement[start_idx].isspace():
            start_idx += 1
        
        # Extract table name (until space or parenthesis)
        end_idx = start_idx
        while end_idx < len(statement) and statement[end_idx] not in (' ', '\t', '\n', '('):
            end_idx += 1
        
        table_name = statement[start_idx:end_idx].lower().strip('`"')
        return table_name
    
    def get_table_priority(statement: str) -> int:
        """Get priority based on table dependency order"""
        table_name = get_table_name(statement)
        try:
            return table_order.index(table_name)
        except ValueError:
            # Unknown tables go at the end
            return len(table_order)
    
    # Sort by table priority
    sorted_statements = sorted(insert_statements, key=get_table_priority)
    
    return sorted_statements


@router.get("/health")
def health_check():
    """Health check endpoint"""
    return {"status": "ok", "message": "Nutrition Analytics API is running"}
